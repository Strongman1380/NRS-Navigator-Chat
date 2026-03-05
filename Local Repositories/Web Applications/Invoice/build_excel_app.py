#!/usr/bin/env python3
"""Child Welfare Service Provider Travel Log System"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

PINK = "FF99CC"
YELLOW = "FFFF00"
BLUE = "333399"

wb = Workbook()
wb.remove(wb.active)

# === ReferenceData Sheet ===
ws = wb.create_sheet("ReferenceData", 0)

# Service Types header
for c, val in enumerate(["Service Type", "Service Code"], 1):
    cell = ws.cell(1, c, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")

services = ["Travel-PT", "FS-OH", "Travel-FS-OH", "FS-IH", "Travel-FS-IH", 
            "DT-Urine 12 panel", "DT-Breath Alcohol", "DT-Oral Fluid", 
            "DT-Patch", "DT-Sweat", "Intake Assessment", "Case Management"]
codes = ["9789", "7443", "2178", "7171", "2178", "4509", "4511", "4512", "4513", "4514", "5100", "5101"]
for r, (s, c) in enumerate(zip(services, codes), 2):
    ws.cell(r, 1, s)
    ws.cell(r, 2, c)

# Staff header
for c, val in enumerate(["Staff Name", "Initials"], 1):
    cell = ws.cell(1, c+4, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")

staff = [("Sheila Decker", "SD"), ("Pamela Spier-Edmond", "PSE"), ("Amanda Weinman", "AW"), 
         ("Kris Zeilinger", "KZ"), ("Kyle Fox", "KF"), ("Onyx Sperry", "OS")]
for r, (name, init) in enumerate(staff, 2):
    ws.cell(r, 5, name)
    ws.cell(r, 6, init)

# Clients header
for c, val in enumerate(["Client Name", "Master Case #", "Status"], 1):
    cell = ws.cell(1, c+7, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")

clients = [("Boruch, Ashley", "604239", "Active"), ("Connelly, Cortney", "485231", "Active"),
           ("Devor, Christina", "190820", "Active"), ("Durflinger, Jamie", "521842", "Active"),
           ("Edmondson, Kimberly", "512893", "Active"), ("Farr, Rebecca", "598234", "Active"),
           ("Gomez, Maria", "567891", "Active"), ("Hernandez, Carlos", "534567", "Active"),
           ("Johnson, Sarah", "623456", "Active"), ("Kowalski, Michael", "589012", "Active"),
           ("Lee, David", "645678", "Active"), ("Martinez, Jennifer", "501234", "Active")]
for r, (name, mc, status) in enumerate(clients, 2):
    ws.cell(r, 8, name)
    ws.cell(r, 9, mc)
    ws.cell(r, 10, status)

# Locations header
for c, val in enumerate(["Location Name", "Address"], 1):
    cell = ws.cell(1, c+11, val)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")

locations = [("Epworth Village (EFR)", "2006 Court St Beatrice, NE"),
             ("Epworth Village (EFS)", "1800 Scott St. Beatrice, NE"),
             ("Boruch Home Norfolk", "308 W Prospect Ave Norfolk, NE"),
             ("Devor Home Beatrice", "525 N 5th St Beatrice, NE"),
             ("Durflinger Home Omaha", "8901 Maple St Omaha, NE"),
             ("Court House Beatrice", "400 Court St Beatrice, NE"),
             ("DCFS Office Lincoln", "301 Centennial Mall Lincoln, NE")]
for r, (name, addr) in enumerate(locations, 2):
    ws.cell(r, 12, name)
    ws.cell(r, 13, addr)

ws.column_dimensions['A'].width = 18
ws.column_dimensions['H'].width = 22
ws.column_dimensions['I'].width = 14
ws.column_dimensions['L'].width = 22
ws.column_dimensions['M'].width = 40
ws.freeze_panes = 'A2'

# === TravelLogs Sheet ===
ws = wb.create_sheet("TravelLogs", 1)
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
header_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ["Entry #", "Sub-Entry", "Index", "Service Type", "Service Notes", "Travel Comments",
           "Client Name", "Master Case #", "Date", "Time In", "Time Out",
           "Drive From", "Drive To", "Miles", "Staff Name",
           "Transport Hrs", "Total Miles", "Billable Hrs", "Indirect Hrs",
           "Drive Total", "Daily Total", "Notes", "Check"]

for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Data validations (using string type for compatibility)
dv = DataValidation(type="list", formula1="=ReferenceData!$A$2:$A$13", allow_blank=True)
dv.add("D2:D1000")
ws.add_data_validation(dv)

dv = DataValidation(type="list", formula1="=ReferenceData!$H$2:$H$13", allow_blank=True)
dv.add("G2:G1000")
ws.add_data_validation(dv)

dv = DataValidation(type="list", formula1="=ReferenceData!$E$2:$E$7", allow_blank=True)
dv.add("O2:O1000")
ws.add_data_validation(dv)

# Sample data
entries = [
    (1, 0, "Travel-PT", "Devor, Christina", 45646, 1, "Onyx Sperry", "325 Eastridge Beatrice", "2006 Court St Beatrice"),
    (1, 1, "FS-IH", "Devor, Christina", 45646, 0, "Onyx Sperry", "", ""),
    (2, 0, "Travel-PT", "Boruch, Ashley", 45646, 2, "Sheila Decker", "Epworth Village (EFR)", "Boruch Home Norfolk"),
    (3, 0, "FS-OH", "Connelly, Cortney", 45647, 0, "Pamela Spier-Edmond", "", ""),
    (4, 0, "Travel-FS-IH", "Durflinger, Jamie", 45647, 15, "Amanda Weinman", "Epworth Village (EFR)", "Durflinger Home Omaha"),
    (5, 0, "DT-Urine 12 panel", "Johnson, Sarah", 45648, 0, "Kris Zeilinger", "", ""),
    (6, 0, "Travel-PT", "Edmondson, Kimberly", 45648, 8, "Kyle Fox", "Epworth Village (EFS)", "Edmondson Home Grand Island"),
    (7, 0, "FS-IH", "Farr, Rebecca", 45649, 0, "Onyx Sperry", "", ""),
    (8, 0, "Travel-FS-OH", "Gomez, Maria", 45649, 12, "Sheila Decker", "Epworth Village (EFR)", "Gomez Home Fremont"),
    (9, 0, "Travel-PT", "Hernandez, Carlos", 45650, 25, "Pamela Spier-Edmond", "Epworth Village (EFR)", "Hernandez Home Norfolk"),
    (10, 0, "FS-OH", "Johnson, Sarah", 45650, 0, "Amanda Weinman", "", ""),
]

for r, (entry, sub, service, client, date_val, miles, staff_name, from_addr, to_addr) in enumerate(entries, 2):
    ws.cell(r, 1, entry)
    ws.cell(r, 2, sub)
    ws.cell(r, 3, f'=IF(B{r}=0,A{r},A{r}+B{r}/100)')
    ws.cell(r, 4, service)
    ws.cell(r, 5, "")
    ws.cell(r, 6, "")
    ws.cell(r, 7, client)
    ws.cell(r, 8, f'=IFERROR(VLOOKUP(G{r},ReferenceData!$H$2:$I$13,2,FALSE),"")')
    ws.cell(r, 9, datetime(2026, 1, 1) + timedelta(days=date_val - 45646))
    ws.cell(r, 10, 0.3333)
    ws.cell(r, 11, 0.5 if "Travel" in service else 0.6667)
    ws.cell(r, 12, from_addr)
    ws.cell(r, 13, to_addr)
    ws.cell(r, 14, miles)
    ws.cell(r, 15, staff_name)
    ws.cell(r, 16, f'=IF(OR(D{r}="Travel-PT",D{r}="Travel-FS-OH",D{r}="Travel-FS-IH"),IF(OR(J{r}="",K{r}=""),0,(K{r}-J{r})*24),0)')
    ws.cell(r, 17, f'=SUM($N$2:N{r})')
    ws.cell(r, 18, f'=IF(OR(D{r}="FS-OH",D{r}="FS-IH",LEFT(D{r},2)="DT"),IF(OR(J{r}="",K{r}=""),0,(K{r}-J{r})*24),0)')
    ws.cell(r, 19, f'=IF(OR(D{r}="FS-OH",D{r}="FS-IH"),IF(OR(J{r}="",K{r}=""),0,(K{r}-J{r})*24),0)')
    ws.cell(r, 20, f'=N{r}')
    ws.cell(r, 21, f'=IF(N{r}>0,T{r},R{r})')
    ws.cell(r, 22, "")
    ws.cell(r, 23, f'=IF(AND(D{r}<>"",G{r}<>"",O{r}<>"",N{r}>=0),"OK","CHECK")')
    
    for c in [3, 8, 16, 17, 18, 19, 20, 21, 23]:
        cell = ws.cell(r, c)
        cell.fill = yellow_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    ws.cell(r, 18).font = Font(bold=True, color=BLUE)

# Column widths
for col, w in [('A',10),('B',10),('C',10),('D',18),('E',25),('F',30),('G',22),('H',14),
               ('I',12),('J',12),('K',12),('L',35),('M',35),('N',10),('O',18),('P',14),
               ('Q',14),('R',14),('S',14),('T',12),('U',12),('V',25),('W',14)]:
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A2'
ws.print_area = 'A1:W60'
ws.page_setup.orientation = 'landscape'

# === Dashboard Sheet ===
ws = wb.create_sheet("Dashboard", 2)
ws.cell(1, 1, "Child Welfare Service Provider Travel Log Dashboard").font = Font(bold=True, size=14)
ws.merge_cells('A1:G1')

metrics = [("Total Entries:", '=COUNTA(TravelLogs!A:A)-1'),
           ("Total Hours:", '=SUM(TravelLogs!R:R)'),
           ("Total Miles:", '=SUM(TravelLogs!N:N)'),
           ("Avg Hours/Trip:", '=IF(B4>0,B5/B4,0)'),
           ("Est. Mileage Cost:", '=B6*0.67')]

for r, (label, formula) in enumerate(metrics, 3):
    ws.cell(r, 1, label)
    ws.cell(r, 2, formula)

ws.cell(3, 2).number_format = '0'
ws.cell(4, 2).number_format = '0.00'
ws.cell(5, 2).number_format = '0'
ws.cell(6, 2).number_format = '0.00'
ws.cell(7, 2).number_format = '$0.00'

ws.cell(9, 1, "STAFF SUMMARY").font = Font(bold=True, size=12)
ws.cell(10, 1, "Staff Name").font = Font(bold=True)
ws.cell(10, 2, "Hours").font = Font(bold=True)
ws.cell(10, 3, "Miles").font = Font(bold=True)
ws.cell(10, 4, "Trips").font = Font(bold=True)

staff_list = ["Sheila Decker", "Pamela Spier-Edmond", "Amanda Weinman", "Kris Zeilinger", "Kyle Fox", "Onyx Sperry"]
for r, sname in enumerate(staff_list, 11):
    ws.cell(r, 1, sname)
    ws.cell(r, 2, f'=SUMIF(TravelLogs!O:O,"{sname}",TravelLogs!R:R)')
    ws.cell(r, 3, f'=SUMIF(TravelLogs!O:O,"{sname}",TravelLogs!N:N)')
    ws.cell(r, 4, f'=COUNTIF(TravelLogs!O:O,"{sname}")')

# === Instructions Sheet ===
ws = wb.create_sheet("Instructions", 3)
ws.cell(1, 1, "Travel Log System Instructions").font = Font(bold=True, size=14)
ws.merge_cells('A1:G1')

instructions = [
    "", "SHEET OVERVIEW:",
    "- TravelLogs - Main data entry sheet for logging all travel and service entries",
    "- ReferenceData - Contains all dropdown lists (service types, clients, staff, locations)",
    "- Dashboard - Summary view showing totals by staff, hours, miles, and costs",
    "", "DATA ENTRY:",
    "1. Select Service Type from dropdown (column D)",
    "2. Select Client Name from dropdown (column G) - Master Case # auto-populates",
    "3. Enter Date, Time In, Time Out",
    "4. Enter Miles driven",
    "5. Select Staff Name from dropdown (column O)",
    "", "PROTECTED COLUMNS (Yellow):",
    "- Columns C, H, P-U, W contain formulas - DO NOT EDIT",
    "", "COLOR CODING:",
    "- Pink headers = Input required",
    "- Yellow cells = Calculated (formula-based, do not edit)",
    "- Blue text = Billable hours",
    "- OK = Valid entry",
    "- CHECK = Entry needs review",
]

for r, line in enumerate(instructions, 3):
    ws.cell(r, 1, line)

ws.column_dimensions['A'].width = 70

# Save
wb.save('TravelLogSystem.xlsx')
print("SUCCESS: Created TravelLogSystem.xlsx")
